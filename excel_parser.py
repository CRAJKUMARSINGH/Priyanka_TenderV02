import pandas as pd
import logging
from typing import Dict, Any, Optional, List
import re
from io import BytesIO
import openpyxl

logger = logging.getLogger(__name__)

class ExcelParser:
    """Enhanced Excel parser with improved error handling and format detection"""
    
    def __init__(self):
        self.required_columns = ['nit_number', 'work_name', 'estimated_cost']
        self.optional_columns = ['schedule_amount', 'earnest_money', 'time_of_completion', 'ee_name', 'date']
        self.bidder_pattern = re.compile(r'bidder\s*(\d+)\s*(name|percentage|contact)', re.IGNORECASE)
    
    def parse_excel(self, uploaded_file) -> Optional[Dict[str, Any]]:
        """Parse Excel file with enhanced format detection and error handling"""
        try:
            # Read file content
            file_content = uploaded_file.read()
            uploaded_file.seek(0)  # Reset file pointer
            
            # Try different parsing methods in order of preference
            parsing_methods = [
                self._parse_nit_multiple_works_format,  # New method for NIT with multiple works
                self._parse_standard_format,
                self._parse_vertical_format,
                self._parse_mixed_format,
                self._parse_any_format
            ]
            
            for method in parsing_methods:
                try:
                    logger.info(f"Trying parsing method: {method.__name__}")
                    uploaded_file.seek(0)  # Reset file pointer
                    result = method(uploaded_file)
                    if result and self._validate_parsed_data(result):
                        logger.info(f"Successfully parsed using {method.__name__}")
                        return result
                except Exception as e:
                    logger.warning(f"Method {method.__name__} failed: {str(e)}")
                    continue
            
            logger.error("All parsing methods failed")
            return None
            
        except Exception as e:
            logger.error(f"Error parsing Excel file: {str(e)}")
            return None
    
    def _parse_nit_multiple_works_format(self, uploaded_file) -> Optional[Dict[str, Any]]:
        """Parse NIT format with multiple works in rows"""
        try:
            workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
            sheet = workbook.active
            
            # Extract header information (rows 1-4)
            nit_number = None
            dates = {}
            
            # Parse header information
            for row in range(1, 5):
                cell_label = sheet.cell(row=row, column=1).value
                cell_value = sheet.cell(row=row, column=3).value
                
                if cell_label and cell_value:
                    label_str = str(cell_label).lower()
                    if 'nit' in label_str and 'number' in label_str:
                        nit_number = str(cell_value)
                    elif 'calling' in label_str:
                        dates['calling_date'] = str(cell_value)
                    elif 'receipt' in label_str:
                        dates['receipt_date'] = str(cell_value)
                    elif 'opening' in label_str:
                        dates['opening_date'] = str(cell_value)
            
            # Check if row 5 contains column headers
            headers_row = 5
            headers = []
            for col in range(1, sheet.max_column + 1):
                header = sheet.cell(row=headers_row, column=col).value
                headers.append(str(header).lower().strip() if header else f"col_{col}")
            
            # Parse works data starting from row 6
            works = []
            for row in range(6, sheet.max_row + 1):
                work_data = {}
                row_has_data = False
                
                for col, header in enumerate(headers, 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value is not None and str(cell_value).strip():
                        row_has_data = True
                        
                        # Map headers to standard field names
                        if 'item' in header or 'no' in header:
                            work_data['item_number'] = str(cell_value).strip()
                        elif 'work' in header or 'name' in header:
                            work_data['work_name'] = str(cell_value).strip()
                        elif 'estimated' in header or 'cost' in header:
                            # Convert from lacs to rupees
                            try:
                                cost_in_lacs = float(cell_value)
                                work_data['estimated_cost'] = cost_in_lacs * 100000  # Convert lacs to rupees
                            except (ValueError, TypeError):
                                work_data['estimated_cost'] = cell_value
                        elif 'schedule' in header or 'g-schedule' in header:
                            work_data['schedule_amount'] = self._clean_numeric_value(cell_value)
                        elif 'completion' in header or 'month' in header:
                            work_data['time_of_completion'] = self._clean_numeric_value(cell_value)
                        elif 'earnest' in header or 'money' in header:
                            work_data['earnest_money'] = self._clean_numeric_value(cell_value)
                
                if row_has_data and work_data:
                    # Add common NIT information to each work
                    work_data['nit_number'] = nit_number
                    work_data['date'] = dates.get('opening_date', dates.get('receipt_date', ''))
                    works.append(work_data)
            
            workbook.close()
            
            if not works:
                return None
            
            # Return structure for multiple works
            result = {
                'nit_number': nit_number,
                'multiple_works': True,
                'works_count': len(works),
                'works': works,
                'dates': dates
            }
            
            logger.info(f"Successfully parsed NIT with {len(works)} works")
            return result
            
        except Exception as e:
            logger.error(f"Error in NIT multiple works parsing: {str(e)}")
            return None
    
    def _clean_numeric_value(self, value):
        """Clean and convert numeric values"""
        if value is None:
            return 0
        try:
            if isinstance(value, (int, float)):
                return float(value)
            # Remove any non-numeric characters and convert
            cleaned = re.sub(r'[^\d.]', '', str(value))
            return float(cleaned) if cleaned else 0
        except:
            return 0
    
    def _parse_standard_format(self, uploaded_file) -> Optional[Dict[str, Any]]:
        """Parse standard horizontal format Excel file"""
        try:
            # Try reading with openpyxl first
            workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
            sheet = workbook.active
            
            # Convert to pandas DataFrame
            data = sheet.values
            cols = next(data)[0:] if data else []
            df = pd.DataFrame(data, columns=cols)
            
            # Clean column names
            df.columns = [str(col).lower().strip().replace(' ', '_') for col in df.columns]
            
            # Remove empty rows
            df = df.dropna(how='all')
            
            if df.empty:
                return None
            
            # Extract data from first row
            extracted_data = {}
            
            # Map common column variations
            column_mapping = {
                'nit_number': ['nit_number', 'nit_no', 'tender_number', 'tender_no'],
                'work_name': ['work_name', 'work_description', 'description', 'work'],
                'estimated_cost': ['estimated_cost', 'estimate', 'cost', 'amount'],
                'schedule_amount': ['schedule_amount', 'schedule', 'sch_amount'],
                'earnest_money': ['earnest_money', 'em', 'security_deposit'],
                'time_of_completion': ['time_of_completion', 'completion_time', 'duration', 'months'],
                'ee_name': ['ee_name', 'executive_engineer', 'engineer_name', 'ee'],
                'date': ['date', 'tender_date', 'submission_date']
            }
            
            # Extract basic data
            for key, variations in column_mapping.items():
                for variation in variations:
                    if variation in df.columns:
                        value = df[variation].iloc[0]
                        if pd.notna(value):
                            extracted_data[key] = self._clean_value(value, key)
                        break
            
            # Extract bidder information
            bidders = self._extract_bidders_horizontal(df)
            if bidders:
                extracted_data['bidders'] = bidders
            
            return extracted_data if extracted_data else None
            
        except Exception as e:
            logger.error(f"Error in standard format parsing: {str(e)}")
            return None
    
    def _parse_vertical_format(self, uploaded_file) -> Optional[Dict[str, Any]]:
        """Parse vertical format where data is in key-value pairs"""
        try:
            df = pd.read_excel(uploaded_file, header=None)
            
            if df.empty or df.shape[1] < 2:
                return None
            
            extracted_data = {}
            
            # Look for key-value pairs in first two columns
            for idx, row in df.iterrows():
                if pd.isna(row.iloc[0]) or pd.isna(row.iloc[1]):
                    continue
                
                key = str(row.iloc[0]).lower().strip()
                value = row.iloc[1]
                
                # Map keys to standard names
                if 'nit' in key or 'tender' in key:
                    extracted_data['nit_number'] = self._clean_value(value, 'nit_number')
                elif 'work' in key or 'description' in key:
                    extracted_data['work_name'] = self._clean_value(value, 'work_name')
                elif 'estimate' in key or 'cost' in key:
                    extracted_data['estimated_cost'] = self._clean_value(value, 'estimated_cost')
                elif 'schedule' in key:
                    extracted_data['schedule_amount'] = self._clean_value(value, 'schedule_amount')
                elif 'earnest' in key or 'security' in key:
                    extracted_data['earnest_money'] = self._clean_value(value, 'earnest_money')
                elif 'completion' in key or 'duration' in key or 'time' in key:
                    extracted_data['time_of_completion'] = self._clean_value(value, 'time_of_completion')
                elif 'engineer' in key or 'ee' in key:
                    extracted_data['ee_name'] = self._clean_value(value, 'ee_name')
                elif 'date' in key:
                    extracted_data['date'] = self._clean_value(value, 'date')
            
            return extracted_data if extracted_data else None
            
        except Exception as e:
            logger.error(f"Error in vertical format parsing: {str(e)}")
            return None
    
    def _parse_mixed_format(self, uploaded_file) -> Optional[Dict[str, Any]]:
        """Parse mixed format with multiple sheets or sections"""
        try:
            # Try reading all sheets
            all_sheets = pd.read_excel(uploaded_file, sheet_name=None)
            
            extracted_data = {}
            
            for sheet_name, df in all_sheets.items():
                if df.empty:
                    continue
                
                # Try standard parsing on this sheet
                sheet_data = self._extract_from_dataframe(df)
                if sheet_data:
                    extracted_data.update(sheet_data)
                
                # Look for bidder information in this sheet
                bidders = self._extract_bidders_from_sheet(df)
                if bidders:
                    if 'bidders' not in extracted_data:
                        extracted_data['bidders'] = []
                    extracted_data['bidders'].extend(bidders)
            
            return extracted_data if extracted_data else None
            
        except Exception as e:
            logger.error(f"Error in mixed format parsing: {str(e)}")
            return None
    
    def _parse_any_format(self, uploaded_file) -> Optional[Dict[str, Any]]:
        """Last resort parser that tries to extract any recognizable data"""
        try:
            # Read entire workbook
            workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
            extracted_data = {}
            
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    if not row or all(cell is None for cell in row):
                        continue
                    
                    row_text = ' '.join([str(cell) for cell in row if cell is not None])
                    
                    # Use regex patterns to extract data
                    patterns = {
                        'nit_number': r'(?:nit|tender)[\s:]*(\d+\/\d{4}-\d{2})',
                        'estimated_cost': r'(?:estimate|cost)[\s:]*(?:rs\.?|\₹)?\s*(\d+(?:,\d+)*(?:\.\d+)?)',
                        'earnest_money': r'(?:earnest|security)[\s:]*(?:rs\.?|\₹)?\s*(\d+(?:,\d+)*(?:\.\d+)?)',
                        'time_of_completion': r'(?:completion|duration)[\s:]*(\d+)\s*(?:months?)',
                    }
                    
                    for key, pattern in patterns.items():
                        if key not in extracted_data:
                            match = re.search(pattern, row_text, re.IGNORECASE)
                            if match:
                                extracted_data[key] = self._clean_value(match.group(1), key)
                    
                    # Look for work name (usually longer text)
                    if 'work_name' not in extracted_data:
                        for cell in row:
                            if cell and isinstance(cell, str) and len(cell) > 20:
                                if any(word in cell.lower() for word in ['work', 'construction', 'repair', 'maintenance']):
                                    extracted_data['work_name'] = cell.strip()
                                    break
            
            return extracted_data if extracted_data else None
            
        except Exception as e:
            logger.error(f"Error in any format parsing: {str(e)}")
            return None
    
    def _extract_from_dataframe(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Extract data from a DataFrame using various heuristics"""
        extracted_data = {}
        
        try:
            # Clean column names
            df.columns = [str(col).lower().strip().replace(' ', '_') for col in df.columns]
            
            # Try to find data in first non-empty row
            for idx, row in df.iterrows():
                if row.notna().any():
                    break
            else:
                return {}
            
            # Column mapping
            column_mapping = {
                'nit_number': ['nit_number', 'nit_no', 'tender_number', 'tender_no'],
                'work_name': ['work_name', 'work_description', 'description', 'work'],
                'estimated_cost': ['estimated_cost', 'estimate', 'cost', 'amount'],
                'schedule_amount': ['schedule_amount', 'schedule', 'sch_amount'],
                'earnest_money': ['earnest_money', 'em', 'security_deposit'],
                'time_of_completion': ['time_of_completion', 'completion_time', 'duration', 'months'],
                'ee_name': ['ee_name', 'executive_engineer', 'engineer_name', 'ee'],
                'date': ['date', 'tender_date', 'submission_date']
            }
            
            for key, variations in column_mapping.items():
                for variation in variations:
                    if variation in df.columns:
                        value = row[variation]
                        if pd.notna(value):
                            extracted_data[key] = self._clean_value(value, key)
                        break
            
            return extracted_data
            
        except Exception as e:
            logger.error(f"Error extracting from DataFrame: {str(e)}")
            return {}
    
    def _extract_bidders_horizontal(self, df: pd.DataFrame) -> List[Dict[str, Any]]:
        """Extract bidders from horizontal format"""
        bidders = []
        
        try:
            # Look for bidder columns
            bidder_columns = {}
            
            for col in df.columns:
                match = self.bidder_pattern.search(str(col))
                if match:
                    bidder_num = int(match.group(1))
                    field_type = match.group(2).lower()
                    
                    if bidder_num not in bidder_columns:
                        bidder_columns[bidder_num] = {}
                    
                    bidder_columns[bidder_num][field_type] = col
            
            # Extract bidder data
            for bidder_num in sorted(bidder_columns.keys()):
                bidder_cols = bidder_columns[bidder_num]
                bidder_data = {}
                
                if 'name' in bidder_cols:
                    name_value = df[bidder_cols['name']].iloc[0]
                    if pd.notna(name_value):
                        bidder_data['name'] = str(name_value).strip()
                
                if 'percentage' in bidder_cols:
                    pct_value = df[bidder_cols['percentage']].iloc[0]
                    if pd.notna(pct_value):
                        bidder_data['percentage'] = self._clean_percentage(pct_value)
                
                if 'contact' in bidder_cols:
                    contact_value = df[bidder_cols['contact']].iloc[0]
                    if pd.notna(contact_value):
                        bidder_data['contact'] = str(contact_value).strip()
                
                if bidder_data and 'name' in bidder_data:
                    bidders.append(bidder_data)
            
            return bidders
            
        except Exception as e:
            logger.error(f"Error extracting bidders horizontally: {str(e)}")
            return []
    
    def _extract_bidders_from_sheet(self, df: pd.DataFrame) -> List[Dict[str, Any]]:
        """Extract bidders from any sheet format"""
        bidders = []
        
        try:
            # Look for columns that might contain bidder names
            name_columns = []
            percentage_columns = []
            
            for col in df.columns:
                col_str = str(col).lower()
                if any(word in col_str for word in ['name', 'bidder', 'contractor', 'company']):
                    name_columns.append(col)
                elif any(word in col_str for word in ['percentage', 'percent', '%', 'rate']):
                    percentage_columns.append(col)
            
            if name_columns:
                name_col = name_columns[0]
                pct_col = percentage_columns[0] if percentage_columns else None
                
                for idx, row in df.iterrows():
                    name = row[name_col]
                    if pd.notna(name) and str(name).strip():
                        bidder_data = {'name': str(name).strip()}
                        
                        if pct_col and pd.notna(row[pct_col]):
                            bidder_data['percentage'] = self._clean_percentage(row[pct_col])
                        
                        bidders.append(bidder_data)
            
            return bidders
            
        except Exception as e:
            logger.error(f"Error extracting bidders from sheet: {str(e)}")
            return []
    
    def _clean_value(self, value, field_type: str):
        """Clean and format values based on field type"""
        if pd.isna(value):
            return None
        
        try:
            if field_type in ['estimated_cost', 'schedule_amount', 'earnest_money']:
                # Clean numeric values
                if isinstance(value, str):
                    # Remove currency symbols and formatting
                    cleaned = re.sub(r'[₹,\s]', '', value)
                    return float(cleaned)
                return float(value)
            
            elif field_type == 'time_of_completion':
                if isinstance(value, str):
                    # Extract number from string
                    match = re.search(r'(\d+)', value)
                    if match:
                        return int(match.group(1))
                return int(float(value))
            
            elif field_type == 'percentage':
                return self._clean_percentage(value)
            
            elif field_type == 'date':
                if isinstance(value, str):
                    return value.strip()
                return str(value)
            
            else:
                # Text fields
                return str(value).strip()
                
        except (ValueError, TypeError) as e:
            logger.warning(f"Could not clean value {value} for field {field_type}: {str(e)}")
            return str(value) if value else None
    
    def _clean_percentage(self, value) -> float:
        """Clean percentage values"""
        try:
            if isinstance(value, str):
                # Remove % symbol and whitespace
                cleaned = value.replace('%', '').strip()
                return float(cleaned)
            return float(value)
        except (ValueError, TypeError):
            logger.warning(f"Could not parse percentage: {value}")
            return 0.0
    
    def _validate_parsed_data(self, data: Dict[str, Any]) -> bool:
        """Validate that parsed data contains minimum required information"""
        if not data:
            return False
        
        # Check for multiple works format
        if data.get('multiple_works'):
            # Validate multiple works structure
            if 'works' not in data or not data['works']:
                logger.warning("Multiple works format detected but no works found")
                return False
            
            # Validate each work has minimum required fields
            for i, work in enumerate(data['works']):
                if 'work_name' not in work or not work['work_name']:
                    logger.warning(f"Work {i+1} missing work_name")
                    return False
                if 'estimated_cost' not in work or not work['estimated_cost']:
                    logger.warning(f"Work {i+1} missing estimated_cost")
                    return False
            
            logger.info(f"Validation result: True, extracted {len(data['works'])} works from NIT {data.get('nit_number')}")
            return True
        
        # Check for at least one required field for single work
        has_required = any(field in data for field in self.required_columns)
        
        # Check for reasonable values
        if 'estimated_cost' in data:
            cost = data['estimated_cost']
            if not isinstance(cost, (int, float)) or cost <= 0:
                return False
        
        if 'nit_number' in data:
            nit = data['nit_number']
            if not isinstance(nit, str) or len(nit.strip()) < 3:
                return False
        
        logger.info(f"Validation result: {has_required}, extracted fields: {list(data.keys())}")
        return has_required

    def get_sample_format(self) -> Dict[str, Any]:
        """Return sample Excel format for user reference"""
        return {
            "description": "Sample Excel format for tender data",
            "columns": {
                "Required": ["NIT Number", "Work Name", "Estimated Cost"],
                "Optional": ["Schedule Amount", "Earnest Money", "Time of Completion", "EE Name", "Date"],
                "Bidders": ["Bidder 1 Name", "Bidder 1 Percentage", "Bidder 1 Contact", "Bidder 2 Name", "..."]
            },
            "sample_row": {
                "NIT Number": "27/2024-25",
                "Work Name": "Construction of Road",
                "Estimated Cost": 1000000,
                "Schedule Amount": 1000000,
                "Earnest Money": 20000,
                "Time of Completion": 12,
                "EE Name": "John Doe",
                "Date": "2024-01-15",
                "Bidder 1 Name": "ABC Company",
                "Bidder 1 Percentage": -5.5,
                "Bidder 1 Contact": "9876543210"
            }
        }
